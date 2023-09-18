# -*- coding: utf-8 -*-
#
# Copyright (c) 2023, CloudBlue LLC
# All rights reserved.
#
from tempfile import NamedTemporaryFile
from typing import Dict, List

from connect.client import AsyncConnectClient, ClientError
from connect.eaas.core.decorators import router, transformation
from connect.eaas.core.inject.asynchronous import get_installation_client
from connect.eaas.core.responses import RowTransformationResponse
from fastapi import Depends
from openpyxl import load_workbook

from connect_transformations.attachment_lookup.exceptions import AttachmentError
from connect_transformations.attachment_lookup.models import Configuration, StreamAttachment
from connect_transformations.attachment_lookup.utils import validate_attachment_lookup
from connect_transformations.models import Error, ValidationResult
from connect_transformations.utils import is_input_column_nullable


class AttachmentLookupTransformationMixin:

    def preload_attachment_for_lookup(self):
        with self.lock():
            if hasattr(self, 'excel_attachments_data'):
                return

            self.excel_attachments_data = {}

            settings = self.transformation_request['transformation']['settings']
            map_by = [item['attachment_column'] for item in settings['map_by']]
            sheet = settings.get('sheet')
            mapping = [col['from'] for col in settings['mapping']]

            file_url = settings['file']
            input_file = NamedTemporaryFile(suffix='.xlsx')
            wb = None

            try:
                url = file_url.split('/public/v1/')[-1]
                url = url[1:] if url[0] == '/' else url
                content = self.installation_client.get(url)
                input_file.write(content)

                wb = load_workbook(input_file, read_only=True)
                if not sheet:
                    sheet = wb.sheetnames[0]
                ws = wb[sheet]
                lookup_columns = {}
                for row in ws.rows:
                    if not lookup_columns:
                        lookup_columns = {
                            cell.value: cell.column - 1
                            for cell in row
                            if cell.value in map_by or cell.value in mapping
                        }
                    else:
                        key = ', '.join([str(row[lookup_columns[item]].value) for item in map_by])
                        self.excel_attachments_data[key] = {
                            col: row[lookup_columns[col]].value for col in mapping
                        }
            except ClientError as e:
                raise AttachmentError(f'Error during downloading attachment: {e}')
            except (KeyError, ValueError) as e:
                raise AttachmentError(f'Invalid column: {e}')
            except Exception as e:
                raise e
            finally:
                if wb:  # pragma: no branch
                    wb.close()
                input_file.close()

    @transformation(
        name='Lookup data from Excel file attached to stream',
        description=(
            'This transformation function allows you to populate data'
            ' from the attached Excel file by matching input columns values'
            ' with the attached table values.'
        ),
        edit_dialog_ui='/static/transformations/attachment_lookup.html',
    )
    def attachment_lookup(
        self,
        row: Dict,
    ):
        try:
            self.preload_attachment_for_lookup()
        except Exception as e:
            return RowTransformationResponse.fail(output=str(e))
        trfn_settings = self.transformation_request['transformation']['settings']
        map_by_from = [item['input_column'] for item in trfn_settings['map_by']]
        input_columns = self.transformation_request['transformation']['columns']['input']

        for item in map_by_from:
            if is_input_column_nullable(
                input_columns,
                item,
            ) and not row[item]:
                return RowTransformationResponse.skip()

        key = ', '.join([str(row[item]) for item in map_by_from])
        attachment_row = self.excel_attachments_data.get(key)
        if not attachment_row:
            return RowTransformationResponse.skip()

        return RowTransformationResponse.done({
            mapping['to']: attachment_row[mapping['from']]
            for mapping in trfn_settings['mapping']
        })


class AttachmentLookupWebAppMixin:

    @router.get(
        '/attachment_lookup/{stream_id}',
        summary='Get a list of Excel files attached to the current stream.',
        response_model=List[StreamAttachment],
        responses={
            400: {'model': Error},
        },
    )
    async def get_excel_attachments(
        self,
        stream_id: str,
        client: AsyncConnectClient = Depends(get_installation_client),
    ):
        query = (
            "((ilike(mime_type,*vnd.openxmlformats-officedocument.spreadsheetml.sheet*)"
            "|ilike(mime_type,*vnd.ms-excel*));"
            "ilike(mime_type,*application*))"
        )
        files = client('media')('folders').collection('streams_attachments')[stream_id].files
        return [
            StreamAttachment(
                id=file['id'],
                name=file['name'],
                file=f'/public/v1/media/folders/streams_attachments/{stream_id}/files/{file["id"]}',
            )
            async for file in files.filter(query)
        ]

    @router.post(
        '/attachment_lookup/validate',
        summary='Validate excel attachment lookup settings',
        response_model=ValidationResult,
        responses={
            400: {'model': Error},
        },
    )
    def validate_attachment_lookup_settings(
        self,
        data: Configuration,
    ):
        return validate_attachment_lookup(data)
